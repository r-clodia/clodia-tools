"""Da documento a Markdown, server-side.

Perché esiste, e perché non basta `_extract_document_text`: quella funzione
serve a FAR LEGGERE un documento a un modello, e per quello il testo piatto va
benissimo. Qui l'obiettivo è diverso — PRODURRE un file `.md` nel topic — e in
quel caso il testo piatto costa caro due volte: perde la struttura (titoli,
articoli, tabelle) e, soprattutto, obbliga il modello a **riemettere in output**
tutto il contenuto per poterlo salvare. Un accordo di 55.000 caratteri sono
~14k token di trascrizione a mano, cioè minuti di generazione per un lavoro che
qui dura un istante — la lentezza segnalata da Davide il 2 set 2026.

Scelta delle dipendenze: `mammoth` (BSD-2) + `markdownify` (MIT), entrambe
**permissive**. Le alternative migliori sul PDF (PyMuPDF/`pymupdf4llm`) sono
AGPL o commerciali Artifex: compatibili con il lato AGPL di questo gateway, non
con l'edizione commerciale del dual-licensing di `LICENSING.md`. Una dipendenza
che vincola una delle due licenze non si aggiunge per comodità.

Conseguenza onesta di quella scelta: **DOCX e XLSX danno Markdown strutturato,
il PDF no.** Da un PDF si ricava testo per riga senza stili, quindi si emette
Markdown minimo (paragrafi separati, marcatore di pagina) e lo si dichiara nel
valore di ritorno con `fidelity`, invece di far credere a chi chiama che la
struttura ci sia. Quando il documento esiste anche in DOCX, convertire quello
è sempre meglio che convertire il PDF che ne è stato stampato.
"""
from __future__ import annotations

import io
import re

#: Come è stato prodotto il Markdown, per chi chiama.
#: `structured` = titoli/liste/tabelle dal formato sorgente;
#: `plain` = righe di testo, nessuno stile recuperabile.
FIDELITY_STRUCTURED = "structured"
FIDELITY_PLAIN = "plain"


def _md_from_docx(data: bytes, revisions: str | None = None) -> tuple[str, dict]:
    """DOCX → Markdown. Ritorna (markdown, statistiche delle revisioni).

    DUE STRADE, e la scelta la fa il documento, non chi chiama.

    Se il file HA revisioni tracciate si usa `docrev`, che le rende esplicite in
    CriticMarkup: `mammoth` include gli inserimenti e scarta le cancellazioni,
    cioè restituisce il documento COME SE le revisioni fossero accettate, senza
    dirlo — su un contratto è la differenza fra ciò che è concordato e ciò che
    una parte ha proposto, e appiattirla in silenzio è il difetto peggiore dei
    due (misurato il 3 set 2026: 7.055 caratteri cancellati spariti senza
    traccia).

    Se NON ne ha, resta `mammoth` + `markdownify`, che mappano gli stili di Word
    (Heading, elenchi, grassetto, tabelle) meglio di quanto faccia un parser
    scritto da noi: là dove non c'è niente da preservare, la libreria vince.

    `revisions` forza la modalità (`inline` | `accepted` | `original`); assente,
    un documento revisionato prende `inline` — chi apre un contratto revisionato
    vuole vedere le revisioni, e il default non deve nascondere informazione.
    """
    from . import docrev

    if revisions or docrev.has_tracked_changes(data):
        return docrev.docx_to_markdown(data, revisions or docrev.MODE_INLINE)

    import mammoth
    from markdownify import markdownify

    html = mammoth.convert_to_html(io.BytesIO(data)).value
    md = markdownify(html, heading_style="ATX", bullets="-")
    # markdownify lascia code di righe vuote fra i blocchi: tre o più newline
    # non hanno significato in Markdown e rendono il file scomodo da leggere.
    md = re.sub(r"\n{3,}", "\n\n", md).strip() + "\n"
    return md, {"revisioni": 0, "inserimenti": 0, "cancellazioni": 0,
                "commenti": 0, "revisori": [], "caratteri": len(md),
                "mode": None}


def _md_from_xlsx(data: bytes) -> str:
    """XLSX → Markdown: un titolo per foglio, una tabella pipe per foglio.

    La prima riga è trattata come intestazione perché è la convenzione di fatto
    dei fogli che si convertono; se non lo è, si legge come una riga di dati in
    grassetto e nessun dato va perso.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[str] = []
    for ws in wb.worksheets:
        out.append(f"## {ws.title}\n")
        righe = [
            ["" if v is None else str(v).replace("|", r"\|").replace("\n", " ")
             for v in row]
            for row in ws.iter_rows(values_only=True)
        ]
        righe = [r for r in righe if any(c.strip() for c in r)]
        if not righe:
            out.append("_(foglio vuoto)_\n")
            continue
        larghezza = max(len(r) for r in righe)
        righe = [r + [""] * (larghezza - len(r)) for r in righe]
        out.append("| " + " | ".join(righe[0]) + " |")
        out.append("|" + "|".join([" --- "] * larghezza) + "|")
        for r in righe[1:]:
            out.append("| " + " | ".join(r) + " |")
        out.append("")
    return "\n".join(out).strip() + "\n"


def _md_from_pdf(data: bytes) -> tuple[str, int]:
    """PDF → Markdown minimo. Ritorna (markdown, n_pagine).

    Senza una libreria che ricostruisca il layout non si distinguono titoli e
    tabelle: si normalizzano i ritorni a capo dentro il paragrafo (i PDF spezzano
    le righe alla larghezza della colonna, non alla fine della frase) e si separa
    ogni pagina con un commento, che è informazione vera e non finge struttura.
    """
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    blocchi: list[str] = []
    for i, pagina in enumerate(reader.pages, start=1):
        testo = pagina.extract_text() or ""
        # Riunisce le righe di un paragrafo: una riga che non chiude una frase e
        # continua in minuscolo è un a-capo tipografico, non un paragrafo nuovo.
        testo = re.sub(r"(?<![.:;!?])\n(?=[a-zà-ù(])", " ", testo)
        testo = re.sub(r"[ \t]+", " ", testo)
        testo = re.sub(r"\n{2,}", "\n\n", testo).strip()
        blocchi.append(f"<!-- pagina {i} -->\n\n{testo}" if testo
                       else f"<!-- pagina {i}: nessun testo estraibile -->")
    return "\n\n".join(blocchi).strip() + "\n", len(reader.pages)


def to_markdown(filename: str, data: bytes,
                revisions: str | None = None) -> tuple[str, int | None, str, dict]:
    """Converte un documento in Markdown.

    Ritorna (markdown, pagine|None, fidelity, revisioni), dove `revisioni` sono
    le statistiche delle revisioni tracciate — vuote per i formati che non le
    hanno. Servono a chi chiama per sapere che un documento È revisionato senza
    doverne rileggere il testo.

    Solleva `ValueError` sulle estensioni che non si sanno convertire, invece di
    ripiegare in silenzio sul testo grezzo: un `.md` che sembra una conversione
    ed è spazzatura decodificata è peggio di un errore, perché nessuno lo
    ricontrolla.
    """
    _vuote = {"revisioni": 0, "inserimenti": 0, "cancellazioni": 0,
              "commenti": 0, "revisori": [], "mode": None}
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "docx":
        md, stats = _md_from_docx(data, revisions)
        return md, None, FIDELITY_STRUCTURED, stats
    if ext in ("xlsx", "xlsm"):
        return _md_from_xlsx(data), None, FIDELITY_STRUCTURED, dict(_vuote)
    if ext == "pdf":
        md, pagine = _md_from_pdf(data)
        return md, pagine, FIDELITY_PLAIN, dict(_vuote)
    if ext in ("md", "markdown", "txt"):
        # Già testo: si normalizza e si passa, così chi chiama non deve sapere
        # in anticipo se il file andava convertito o no.
        return (data.decode("utf-8", errors="replace").strip() + "\n", None,
                FIDELITY_PLAIN, dict(_vuote))
    raise ValueError(
        f"non so convertire '.{ext}' in Markdown (supportati: docx, xlsx, pdf, txt, md)")


def markdown_name(path: str) -> str:
    """Nome del `.md` di destinazione a partire dal path del sorgente.

    Tiene il nome del documento e cambia solo l'estensione: chi guarda i file del
    topic vede la coppia accanto, senza dover ricostruire quale `.md` viene da
    quale allegato.
    """
    base = (path or "").rsplit("/", 1)[-1]
    radice = base.rsplit(".", 1)[0] if "." in base else base
    return f"{radice or 'documento'}.md"


# ── Finestra di lettura ───────────────────────────────────────────────────────
#
# `read_document` consegnava `text[:max_chars]`: sempre dal PRINCIPIO, senza un
# modo di chiedere il resto. Su un contratto di 71.801 caratteri (il dc10 di
# titul-brightnode, 4 set 2026) significa che gli articoli finali — 8.3, 9.1,
# 10, 11, 12, 13 — non erano raggiungibili in nessun modo: non «lenti da
# leggere», irraggiungibili per costruzione. L'agente sapeva solo `truncated:
# true`, che dice che manca qualcosa e non come averlo, e da lì l'unica mossa
# era chiedere a un umano di estrarre i pezzi a mano.
#
# Il taglio va su un confine di RIGA: tagliare a metà carattere di un marcatore
# CriticMarkup (`{++`, `{--`) produrrebbe due finestre in cui una revisione
# sembra aperta e mai chiusa — cioè, su un documento in revisione, un errore di
# lettura che somiglia a un dato.

#: Quanto indietro si può arretrare per trovare un capo a riga prima di
#: rassegnarsi a tagliare a metà. Su testo normale il capo è vicino; il limite
#: serve al caso patologico di una riga lunghissima (una tabella su una riga).
_RITORNO_MAX = 2000


def finestra(text: str, offset: int = 0, max_chars: int = 60000) -> dict:
    """Una finestra del testo, e come chiedere la prossima.

    Ritorna `{text, offset, next_offset, remaining, truncated, window}`.
    `next_offset` è `None` quando non resta niente: chi legge non deve dedurre
    la fine da un confronto fra numeri.
    """
    total = len(text or "")
    off = max(0, int(offset or 0))
    cap = max(1, int(max_chars or 60000))
    if off >= total:
        return {"text": "", "offset": total, "next_offset": None,
                "remaining": 0, "truncated": False, "window": 0}
    fine = min(total, off + cap)
    if fine < total:
        # Arretra all'ultimo capo a riga dentro la finestra, così i marcatori
        # non si spezzano e la finestra dopo comincia da una riga intera.
        taglio = text.rfind("\n", off, fine)
        if taglio > off and (fine - taglio) <= _RITORNO_MAX:
            fine = taglio + 1
    pezzo = text[off:fine]
    resta = total - fine
    return {"text": pezzo, "offset": off,
            "next_offset": fine if resta else None,
            "remaining": resta, "truncated": bool(resta), "window": len(pezzo)}


#: Un'intestazione Markdown (`## Art. 8`) — è così che `docrev` rende i titoli.
_TITOLO = re.compile(r"^(#{1,6})\s+(.+?)\s*$")


def outline(text: str, limite: int = 200) -> list[dict]:
    """I titoli del documento con il loro offset, per saltare al punto giusto.

    Serve perché `next_offset` da solo permette di scorrere in avanti, e chi
    cerca «l'art. 9.1» non vuole scorrere: vuole l'offset di quell'articolo. Con
    l'indice, una lettura mirata costa una finestra invece di quattro.
    """
    fuori: list[dict] = []
    pos = 0
    for riga in (text or "").splitlines(keepends=True):
        m = _TITOLO.match(riga.rstrip("\n"))
        if m:
            fuori.append({"level": len(m.group(1)), "title": m.group(2)[:120],
                          "offset": pos})
            if len(fuori) >= limite:
                break
        pos += len(riga)
    return fuori
